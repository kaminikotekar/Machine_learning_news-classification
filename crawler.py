import requests,bs4,sys,nltk
from nltk.corpus import stopwords
from nltk.tokenize import word_tokenize
import openpyxl
import os

os.chdir("C:\\Users\\ARUN MADHAV\\kaminishivani\\Machine_learning_news classification")
res=requests.get("https://www.ndtv.com/topic/top-50")
res.raise_for_status()
soup=bs4.BeautifulSoup(res.text,'html.parser')
element=soup.select("#news_list p.header.fbld a")
doc_id=1

wb=openpyxl.Workbook()
sheet=wb.get_sheet_by_name('Sheet')
sheet['A1'].value='url_id'
sheet['B1'].value='content'
sheet['C1'].value='class'

stop_words=set(stopwords.words('english'))
newStopWords = set([':',';','[',']','{','}','(',')',',','?','#','@','$','%','^','&','*','!','<','>','/','"',"'",'.','-','+','=','I','""',"''"])
stop_words.update(newStopWords)


def create_dataset(doc_id,li,filtered_sentence,news_type):
    sheet.cell(row=doc_id+1, column=1).value=doc_id
    sheet.cell(row=doc_id+1,column=2).value=' '.join(filtered_sentence)
    sheet.cell(row=doc_id+1,column=3).value=news_type


for each in element:
    li=each.get('href')
    print(li)
    print("\n\n\n")
    if "gadgets" not in li and "sports" not in li and "javascript" not in li:
        resl=requests.get(li)
        soupl=bs4.BeautifulSoup(resl.text,'html.parser')
        content=soupl.select("#ins_storybody")

        for piece in content:
            node=''.join(piece.findAll(text=True))

        word_token =word_tokenize(node)
        filtered_sentence=[]

        for w in word_token:
            if w not in stop_words:
                filtered_sentence.append(w)

        print(filtered_sentence)
        print("\n\n\n")
        create_dataset(doc_id,li,filtered_sentence,"news")
        doc_id+=1

    elif "sports" in li:
        res2=requests.get(li)
        soup2=bs4.BeautifulSoup(res2.text,'html.parser')
        content=soup2.select(".article-detail")

        for piece in content:
            node=''.join(piece.findAll(text=True))

        word_token =word_tokenize(node)
        filtered_sentence=[]

        for w in word_token:
            if w not in stop_words:
                filtered_sentence.append(w)

        print(filtered_sentence)
        print("\nSports\n\n")
        create_dataset(doc_id,li,filtered_sentence,"sports")
        doc_id+=1


    elif "gadgets" in li:
        res3=requests.get(li)
        soup3=bs4.BeautifulSoup(res3.text,'html.parser')
        content=soup3.select("#center_content_div > div")

        for piece in content:
            node=''.join(piece.findAll(text=True))

        word_token =word_tokenize(node)
        filtered_sentence=[]

        for w in word_token:
            if w not in stop_words:
                filtered_sentence.append(w)

        create_dataset(doc_id,li,filtered_sentence,"gadgets")
        print(filtered_sentence)
        print("\n Gadgets\n\n")
        doc_id+=1

    elif "auto" in li:
        res4=requests.get(li)
        soup4=bs4.BeautifulSoup(res4.text,'html.parser')
        content=soup4.select("#content > div.cnb-page.row.row-g_60 > div.cnb-content.grid_17 > div.article.js-photoswipe > div.article__content.mb_40")

        for piece in content:
            node=''.join(piece.findAll(text=True))

        word_token =word_tokenize(node)
        filtered_sentence=[]

        for w in word_token:
            if w not in stop_words:
                filtered_sentence.append(w)

        create_dataset(doc_id,li,filtered_sentence,"Auto")
        print(filtered_sentence)
        print("\n Auto\n\n")
        doc_id+=1

    else:
        continue


wb.save("News_dataset.xlsx")        



 
    
        
        
        
        
        
